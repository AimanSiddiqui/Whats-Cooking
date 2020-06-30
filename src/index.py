# importing the required libraries and modules
from tkinter import *
import json
import re
import string
import unicodedata
from datetime import datetime
import nltk
import numpy as np
from nltk.stem import WordNetLemmatizer
from pip._vendor.pyparsing import unicode
import math
import win32com.client
import openpyxl

np.set_printoptions(threshold=np.inf)
Centroids={}
StopWords = ''
BoW = []
idf_saved = [0]
def import_nltk_data(path):
    # for lemmatizer, nltk_data was downloaded to resource
    # to use that data, we append its path to nltk`s data-path
    if not nltk.data.path.__contains__(path):
        nltk.data.path.append(path)
    return


# ---------------------------------Strip Accents------------------------------------------
# remove all the accents
def strip_accents(text):
    try:
        text = unicode(text, 'utf-8')
    except (TypeError, NameError):  # unicode is a default on python 3
        pass

    text = unicodedata.normalize('NFD', text) \
        .encode('ascii', 'ignore') \
        .decode("utf-8")

    return str(text)

def StopWordsRemoval(ingredients, StopWords):

    ingredients = ingredients.split(" ")
    for i in StopWords:
        if i in ingredients:
            ingredients.remove(i)
    ingredients = " ".join(ingredients)
    ingredients.replace("\t",' ')
    ingredients.replace("    ",' ')
    ingredients.replace("   ",' ')
    ingredients.replace("  ",' ')
    return ingredients
# ---------------------------------LEMMATIZATION---------------------------------------------
# this function performs lemmatization over the string to array convert documents and queries
def Lemmatization(sentence):
    # sent = TextBlob(sentence)
    # tag_dict = {"J": 'a',
    #             "N": 'n',
    #             "V": 'v',
    #             "R": 'r'}
    # words_and_tags = [(w, tag_dict.get(pos[0], 'n')) for w, pos in sent.tags]
    # lemmatized_list = [wd.lemmatize(tag) for wd, tag in words_and_tags]
    # return " ".join(lemmatized_list)
    Lemmatizer = WordNetLemmatizer()
    sentence = Lemmatizer.lemmatize(sentence)
    return sentence


# ---------------------------------NORMALIZATION------------------------------------------
# remove punctuations and spaces,etc i.e this function performs normalizations
def Normalization(doc):
    # doc = re.sub(r"(?<!\w)([A-Za-z])\.", r"\1", doc)
    doc = re.sub("^\d+\s|\s\d+\s|\s\d+$", " ", doc)
    # doc = re.sub(r"[^a-z']", " ", doc)
    # doc = (decontracted(doc))
    Regex = re.compile('[%s]' % re.escape(string.punctuation))
    doc = Regex.sub(' ', doc)
    doc = doc.replace('â€', " ")
    doc = doc.replace("â–", " ")
    doc = doc.replace('”', " ")
    doc = doc.lower()
    doc = strip_accents(doc)
    return doc



def alert_popup(title, message, path,b_message):
    """Generate a pop-up window for special messages."""
    root = Tk()
    root.title(title)
    root.configure(bg="#DB450A")
    w = 600     # popup window width
    h = 300     # popup window height
    sw = root.winfo_screenwidth()
    sh = root.winfo_screenheight()
    x = (sw - w)/2
    y = (sh - h)/2
    root.geometry('%dx%d+%d+%d' % (w, h, x, y))
    m = message
    w = Label(root, text=m, height = 5, font =("Courier New", "12", "bold"), bg="#DB450A")
    w.pack()
    m = path.upper()
    w = Label(root, text=m, height = 3, font =("Courier New", "17", "bold"), bg="#DB450A")
    w.pack()

    b = Button(root, text=b_message, command=root.destroy, width=10)
    b.pack()
    mainloop()
# Examples



def Train():
    global Bow
    global Centroids
    global idf_saved
    if (Centroids):
        alert_popup("Training", "Training already completed!", " ","OK")
    else:

        print("Training!")
        print(datetime.now().strftime("%H:%M:%S") + ": start")
        import_nltk_data("nltk_data/")
        # opening the training file
        with open('C:/Users/Asim Ahmed/Desktop/IR/data/train.json') as f:
            data = json.load(f)
        
        Cuisines = {}
        # Cuisine Structure: Cuisine_name :  count_docs_of_cuisin
        f = open("C:/Users/Asim Ahmed/Desktop/IR/result/BagOfWords.txt", 'w')
        # a = ""
        f2 = open("C:/Users/Asim Ahmed/Desktop/IR/data/StopWords.txt", 'r')
        StopWords = f2.read()
        StopWords = StopWords.split("\n")
        for i in range(0, int(len(data))):
            for j in range(0, len(data[i]['ingredients'])):
            # as data is stored in the form of json object i.e. dictionary hence lemmatizing and normalizing each
            # word of ingredient in it
                data[i]['ingredients'][j] = Normalization(data[i]['ingredients'][j])
                data[i]['ingredients'][j] = StopWordsRemoval(data[i]['ingredients'][j], StopWords)
                data[i]['ingredients'][j] = Lemmatization(data[i]['ingredients'][j])
                # if the word doesnot exist in the Bag of Words inly then is appended in the list of bag of words
                if not (data[i]['ingredients'][j] in BoW):
                    BoW.append(data[i]['ingredients'][j])

                # if the word doesnot exist in the Cuisine dictionary only then is appended in the dictionary of cuisines
                if not (data[i]['cuisine'] in Cuisines.keys()):
                    Cuisines[data[i]['cuisine']] = 1
                else:
                    Cuisines[data[i]['cuisine']] += 1

        print(str(Cuisines.keys()))
        
        f.write(str(BoW))
        f.close()
        # print(json.dumps(data[0], indent = 4, sort_keys=True))
        # print(len(BoW))
        data_len = len(data)
        bow_len = len(BoW)
        print("Bag of Words Length : " , bow_len)

        for i in Cuisines.keys():
            Centroids[i] = [0] * bow_len

        # here tf contains the term frequency of the ingredients in the document.
        # on one dimension we have all the dishIDs saved as keys in the dictionary tf
        # and the value in it contains the frequency of the each of ingredients in the bagwords that are in that dishID
        tf = {}
        # tf = {"id-1": [tf-1, tf-2, tf-3, ...], "id-2": [tf-1, tf-2, ...], ...}

        for i in range(0, int(data_len)):

            doc_id = data[i]['id']
            ingredients = data[i]['ingredients']
            tf[doc_id] = [0] * bow_len
            for j in range(0, bow_len):
                if BoW[j] in ingredients:
                    tf[doc_id][j] += 1
        # BoW.clear()
        # f.close()
        j = 0
        idf_saved = [0] * bow_len
        for i in range(0, bow_len):
            df = 0

            for doc_id in tf.keys():
                if tf.get(doc_id)[i] != 0:
                    df += 1

            j += 1
            idf = float(format(math.log10(data_len / df), '.5f'))
            idf_saved[i] = idf
            for doc_id in tf.keys():
                tf.get(doc_id)[i] *= idf
        
        y = 0
        for doc_id in tf.keys():
            single_cuisine = data[y]['cuisine']
            a = np.array(tf.get(doc_id))
            b = np.array(Centroids[single_cuisine])
            add = np.add(a, b)
            Centroids[single_cuisine] = add
            y += 1

        workbook = openpyxl.Workbook()
    # creates a sheet
        doc_sheet = workbook.active
        for i in range(0,len(BoW)):
          doc_sheet.cell(i+2, 1).value =BoW[i]
          count_j = 2
          for j in  Centroids.keys():
            doc_sheet.cell(1, count_j).value = j
            count_j += 1
          count_j = 2
          for j in  Centroids.keys():
            doc_sheet.cell(i+2, count_j).value = format(Centroids[j][i],'.5f')
            count_j += 1
        # saves the cache file on disk
        workbook.save("C:/Users/Asim Ahmed/Desktop/IR/result/tf-idf.xlsx")

        tf.clear()
        # manually compute cosine similarity
        for i in Centroids.keys():
            Centroids[i] = np.divide(Centroids[i], Cuisines[i] )
        print(datetime.now().strftime("%H:%M:%S") + ": end")
        alert_popup("Training", "Training Completed!", "","OK")

def Test():
    print("Testing!")
    Ingredients = text.get('1.0', END)
    text.delete("1.0","end")
    query = Ingredients.split("\n")
    len_query = len(query)

    for i in range(0,len_query):
        query[i] = Normalization(query[i])
        query[i] = StopWordsRemoval(query[i],StopWords)
        query[i] = Lemmatization(query[i])
    bow_len = len(BoW)
    query_tf = [0] * bow_len
    for j in range(0, bow_len):
        if(BoW[j] in query):
            query_tf[j] = query.count(BoW[j])
        query_tf[j] = query_tf[j] * idf_saved[j]

    query_tf = np.array(query_tf)
    min_dist = -1
    class_cuisine = ''
    for j in Centroids.keys():
        dot = np.dot(Centroids[j],query_tf)
        # print(dot)
        norma = np.linalg.norm(Centroids[j])
        normb = np.linalg.norm(query_tf)
        dist = dot / (norma * normb)
        # dist = np.linalg.norm(Centroids[j]-query_tf)
        print(j , "- ", dist , "- ", norma , "- ", normb)
        if dist > min_dist:
            min_dist = dist
            class_cuisine = j

    print("Class: \n")
    print(class_cuisine)
    


    alert_popup("Classification Answer", "The ingredients result in the following Cuisine:", class_cuisine,"OK")
    print(class_cuisine)


def AddIngredient():
    ingredient = E1.get()
    E1.delete(0, 'end')
    #displaying the result on the screen
    text.insert(INSERT,ingredient +"\n")


main = Tk()




#GUI Making and its steps
main.wm_attributes(None, "white")
bg_color = "#104279"
bt_color = "#FF7E39"
bt_ft_color = "white"

main.title("Cuisine Classification")
main.geometry("1000x600")
main.configure(bg=bg_color)


Image_path = "../data/background.gif"
# background_image= PhotoImage(file=Image_path)
# background_label = Label(main, image=background_image)
# background_label.place(x=0, y=0, relwidth=1, relheight=1)


frame = Frame(main, bg=bg_color)


frame1 = Frame(frame, bg=bg_color)




B1 = Button(frame1, text="Train", bd =1, font = ("Times", "12" ),width = 20, bg= bt_color, fg=bt_ft_color, padx=30, command = Train)
B1.pack()
labelframe = LabelFrame(frame1, text= "    ",height= 10, bg=bg_color)
labelframe.pack()
L1 = Label(frame1, text="Enter Ingredients", font =("Courier New", "15", "bold"), bg=bg_color, fg="white",)
L1.pack()
E1 = Entry(frame1,text="Enter Ingredients", bd =1, font = ("Times", "12" ), width = 50)
E1.pack()
labelframe = LabelFrame(frame1, text= "    ",height= 10, bg=bg_color)
labelframe.pack()
B1 = Button(frame1, text="Add", bd =1, font = ("Times", "12" ),width = 20, bg= bt_color, fg=bt_ft_color, padx=30, command = AddIngredient)
B1.pack()
labelframe = LabelFrame(frame1, text= "    ",height= 10, bg=bg_color)
labelframe.pack()

B1 = Button(frame1, text="Test", bd =1, font = ("Times", "12" ),width = 20, bg= bt_color, fg=bt_ft_color, padx=30, command = Test)
B1.pack()

frame1.pack(side=LEFT)




frame2 = Frame(frame , bd =10, bg=bg_color)
# background_label = Label(frame2, image=background_image)
# background_label.place(x=0, y=0, relwidth=1, relheight=1)

labelframe = LabelFrame(frame2, text= "    ",height= 10, bg=bg_color)
labelframe.pack()

frame1 = Frame(frame2,bd=1, bg="black")

head1 = Text(frame1, bg=bt_color, height=1, width=52,bd=1 ,font =("Consolas", "11", "bold"),fg=bt_ft_color)
head1.bind('<Control-x>', lambda e: 'break')
head1.bind('<Button-3>', lambda e: 'break')
head1.bind('<Button-1>', lambda e: 'break')
head1.pack(side=LEFT)
head1.insert(INSERT, '  INGREDIENTS')

frame1.pack()

scroll = Scrollbar(frame2)
scroll.pack(side=RIGHT, fill=Y)

text = Text(frame2, bg="white", bd=2, height=30,width=50,  yscrollcommand=scroll.set ,wrap="none")
text.bind('<Control-x>', lambda e: 'break')
text.bind('<Button-3>', lambda e: 'break')
text.bind('<Button-1>', lambda e: 'break')
text.pack(expand=True, side= LEFT)
frame2.pack()
frame.pack()


def yview(*args):
        text.yview(*args)
        text1.yview(*args)
        text2.yview(*args)

scroll.config(command=yview)



main.mainloop()

